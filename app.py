import os
import csv
import smtplib
from io import StringIO, BytesIO
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect,
    url_for, flash, Response, send_file
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import CheckConstraint
from dotenv import load_dotenv


# ============================================================
#  Inicialização do app + carregamento do .env
# ============================================================

print(">> RUNNING APP FROM:", __file__)

load_dotenv(dotenv_path=Path(__file__).with_name(".env"), override=True)

app = Flask(__name__)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-key")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///warehouse.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

LOW_STOCK_THRESHOLD = int(os.getenv("LOW_STOCK_THRESHOLD", "3"))


# ============================================================
#  Recarrega .env a cada request (útil no Colab)
# ============================================================

@app.before_request
def _reload_env():
    load_dotenv(dotenv_path=Path(__file__).with_name(".env"), override=True)


# ============================================================
#  Disponibilizar datetime dentro dos templates
# ============================================================

@app.context_processor
def inject_now():
    return {"datetime": datetime}


# ============================================================
#  MODELOS
# ============================================================

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    sku = db.Column(db.String(64), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Integer, nullable=False, default=0)
    min_threshold = db.Column(db.Integer, nullable=False, default=LOW_STOCK_THRESHOLD)
    allow_negative = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    __table_args__ = (
        CheckConstraint('min_threshold >= 0', name='threshold_non_negative'),
    )


class Movement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    delta = db.Column(db.Integer, nullable=False)
    note = db.Column(db.String(300))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    product = db.relationship("Product", backref=db.backref("movements", lazy=True))


# ============================================================
#  FUNÇÃO PARA ENVIAR EMAIL DE BAIXO ESTOQUE
# ============================================================

def send_low_stock_email(product):
    host = os.getenv("SMTP_HOST")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USERNAME")
    pwd = os.getenv("SMTP_PASSWORD")
    use_tls = os.getenv("SMTP_USE_TLS", "1") == "1"
    recipients = [e.strip() for e in os.getenv("ALERT_RECIPIENTS", "").split(",") if e.strip()]
    from_email = os.getenv("FROM_EMAIL", user)

    if not (host and user and pwd and recipients and from_email):
        app.logger.warning("Email not sent — SMTP config incomplete.")
        return False

    msg = EmailMessage()
    msg["Subject"] = f"[Inventory Alert] Low stock for {product.name} (SKU {product.sku})"
    msg["From"] = from_email
    msg["To"] = ", ".join(recipients)

    msg.set_content(
        f"Product: {product.name} (SKU: {product.sku})\n"
        f"Quantity now: {product.quantity}\n"
        f"Threshold: {product.min_threshold}\n"
        f"Time: {datetime.utcnow().isoformat()}Z\n\n"
        "Please reorder or investigate."
    )

    with smtplib.SMTP(host, port) as s:
        if use_tls:
            s.starttls()
        s.login(user, pwd)
        s.send_message(msg)

    return True


# ============================================================
#  HELPERS
# ============================================================

def _get_bool(name, default=False):
    val = request.form.get(name)
    if val is None:
        return default
    return str(val).lower() in ("1", "true", "on", "yes")


# ============================================================
#  ROTAS PRINCIPAIS
# ============================================================

@app.route("/")
def index():
    q = request.args.get("q", "").strip()

    if q:
        products = Product.query.filter(
            (Product.name.ilike(f"%{q}%")) | (Product.sku.ilike(f"%{q}%"))
        ).order_by(Product.name).all()
    else:
        products = Product.query.order_by(Product.name).all()

    low = [p for p in products if p.quantity < p.min_threshold]

    return render_template(
        "products.html",
        products=products,
        low=low,
        q=q,
        threshold_default=LOW_STOCK_THRESHOLD
    )


# ============================================================
#  CRUD - CREATE / UPDATE / DELETE
# ============================================================

@app.route("/product/new", methods=["GET", "POST"])
def product_new():
    if request.method == "POST":
        sku = request.form["sku"].strip()
        name = request.form["name"].strip()
        qty = int(request.form.get("quantity", "0"))
        thr = int(request.form.get("min_threshold", str(LOW_STOCK_THRESHOLD)))
        allow_neg = _get_bool("allow_negative", True)

        if Product.query.filter_by(sku=sku).first():
            flash("SKU already exists.", "error")
            return redirect(url_for("product_new"))

        p = Product(
            sku=sku,
            name=name,
            quantity=qty,
            min_threshold=max(0, thr),
            allow_negative=allow_neg
        )
        db.session.add(p)
        db.session.commit()

        flash("Product created.", "success")
        return redirect(url_for("index"))

    return render_template("product_form.html", product=None, threshold_default=LOW_STOCK_THRESHOLD)


@app.route("/product/<int:pid>/edit", methods=["GET", "POST"])
def product_edit(pid):
    p = Product.query.get_or_404(pid)

    if request.method == "POST":
        p.sku = request.form["sku"].strip()
        p.name = request.form["name"].strip()
        p.quantity = int(request.form.get("quantity", p.quantity))
        p.min_threshold = max(0, int(request.form.get("min_threshold", p.min_threshold)))
        p.allow_negative = _get_bool("allow_negative", p.allow_negative)

        db.session.commit()
        flash("Product updated.", "success")
        return redirect(url_for("index"))

    return render_template("product_form.html", product=p, threshold_default=LOW_STOCK_THRESHOLD)


@app.route("/product/<int:pid>/delete", methods=["POST"])
def product_delete(pid):
    p = Product.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()

    flash("Product deleted.", "success")
    return redirect(url_for("index"))


# ============================================================
#  MOVIMENTAÇÕES
# ============================================================

@app.route("/product/<int:pid>/add", methods=["POST"])
def product_add(pid):
    p = Product.query.get_or_404(pid)
    amount = int(request.form.get("amount", "0"))
    note = request.form.get("note", "").strip()

    if amount <= 0:
        flash("Amount must be positive.", "error")
        return redirect(url_for("index"))

    p.quantity += amount

    db.session.add(Movement(product=p, delta=amount, note=note))
    db.session.commit()

    flash(f"Added {amount} to {p.name}.", "success")
    return redirect(url_for("index"))


@app.route("/product/<int:pid>/remove", methods=["POST"])
def product_remove(pid):
    p = Product.query.get_or_404(pid)
    amount = int(request.form.get("amount", "0"))
    note = request.form.get("note", "").strip()

    if amount <= 0:
        flash("Amount must be positive.", "error")
        return redirect(url_for("index"))

    if not p.allow_negative and amount > p.quantity:
        flash("Este produto não permite estoque negativo.", "error")
        return redirect(url_for("index"))

    p.quantity -= amount

    db.session.add(Movement(product=p, delta=-amount, note=note))
    db.session.commit()

    if p.quantity < p.min_threshold:
        if send_low_stock_email(p):
            flash(f"Low-stock email sent for {p.name}.", "info")
        else:
            flash("Low-stock email NOT sent (check SMTP settings).", "warning")
    else:
        flash(f"Removed {amount} from {p.name}.", "success")

    return redirect(url_for("index"))


# ============================================================
#  RELATÓRIO WEB
# ============================================================

@app.route("/report")
def report():
    products = Product.query.order_by(Product.name).all()
    total_skus = len(products)
    low = [p for p in products if p.quantity < p.min_threshold]

    return render_template(
        "report.html",
        products=products,
        total_skus=total_skus,
        low=low
    )


# ============================================================
#  EXPORTAR CSV
# ============================================================

@app.route("/report/download")
def report_download_csv():
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["SKU", "Name", "Quantity", "Min", "Status", "AllowNegative"])

    products = Product.query.order_by(Product.name).all()

    for p in products:
        status = "LOW" if p.quantity < p.min_threshold else "OK"
        writer.writerow([
            p.sku, p.name, p.quantity, p.min_threshold,
            status, "YES" if p.allow_negative else "NO"
        ])

    output.seek(0)

    filename = f"inventory_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%SZ')}.csv"

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
#  EXPORTAR PDF
# ============================================================

@app.route("/report/download/pdf")
def report_download_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib import colors
    except Exception as e:
        return (
            "ReportLab não está instalado. Rode: <code>pip install reportlab</code><br>"
            f"Erro: {e}", 500
        )

    products = Product.query.order_by(Product.name).all()

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # Cabeçalho
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, height - 20 * mm, "Inventory Report")

    c.setFont("Helvetica", 9)
    c.drawString(20 * mm, height - 26 * mm,
                 f"Generated at: {datetime.utcnow().isoformat()}Z")

    # Tabela
    y = height - 40 * mm
    line_height = 7 * mm

    headers = ["SKU", "Name", "Qty", "Min", "Status", "Neg?"]
    col_x = [12*mm, 50*mm, 130*mm, 150*mm, 170*mm, 188*mm]

    c.setFont("Helvetica-Bold", 10)
    for i, htext in enumerate(headers):
        c.drawString(col_x[i], y, htext)

    y -= line_height
    c.setLineWidth(0.5)
    c.line(12*mm, y + line_height/2, 200*mm, y + line_height/2)

    c.setFont("Helvetica", 10)

    for p in products:
        status = "LOW" if p.quantity < p.min_threshold else "OK"

        if y < 20*mm:
            c.showPage()
            y = height - 20*mm

        c.drawString(col_x[0], y, str(p.sku))
        c.drawString(col_x[1], y, str(p.name)[:42])

        if p.quantity < 0:
            c.setFillColor(colors.red)
        c.drawRightString(col_x[2] + 8, y, str(p.quantity))
        c.setFillColor(colors.black)

        c.drawRightString(col_x[3] + 8, y, str(p.min_threshold))

        c.setFillColor(colors.red if status == "LOW" else colors.green)
        c.drawString(col_x[4], y, status)
        c.setFillColor(colors.black)

        c.drawString(col_x[5], y, "YES" if p.allow_negative else "NO")

        y -= line_height

    c.showPage()
    c.save()

    buf.seek(0)
    filename = f"inventory_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%SZ')}.pdf"

    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )


# ============================================================
#  EXPORTAR XLSX
# ============================================================

@app.route("/report/download/xlsx")
def report_download_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import CellIsRule

    products = Product.query.order_by(Product.name).all()

    # Cria workbook e aba
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Cabeçalhos
    headers = ["SKU", "Name", "Quantity", "Min", "Status", "AllowNegative"]
    ws.append(headers)

    # Deixa o cabeçalho em negrito
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Linhas de dados
    for p in products:
        status = "LOW" if p.quantity < p.min_threshold else "OK"
        allow_neg = "YES" if p.allow_negative else "NO"
        ws.append([
            p.sku,
            p.name,
            p.quantity,
            p.min_threshold,
            status,
            allow_neg,
        ])

    # Ajusta larguras das colunas
    widths = {"A": 18, "B": 40, "C": 12, "D": 10, "E": 10, "F": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Formatação condicional: Status = LOW em vermelho na coluna E
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "E2:E50000",
        CellIsRule(operator="equal", formula=['"LOW"'], fill=red_fill)
    )

    # Formatação condicional: Quantity < 0 em vermelho na coluna C
    ws.conditional_formatting.add(
        "C2:C50000",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
    )

    # Salva em memória e envia
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"inventory_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%SZ')}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# ============================================================
#  ADMIN
# ============================================================

@app.route("/admin/run_stock_check")
def run_stock_check():
    products = Product.query.order_by(Product.name).all()
    count = 0

    for p in products:
        if p.quantity < p.min_threshold and send_low_stock_email(p):
            count += 1

    flash(f"Sent {count} low-stock emails.", "info")
    return redirect(url_for("index"))


@app.route("/admin/debug_smtp")
def debug_smtp():
    keys = [
        "SMTP_HOST", "SMTP_PORT", "SMTP_USERNAME", "SMTP_PASSWORD",
        "FROM_EMAIL", "ALERT_RECIPIENTS", "SMTP_USE_TLS"
    ]

    def mask(value, key):
        if key == "SMTP_PASSWORD" and value:
            return "*" * len(value)
        return value

    vals = {k: mask(os.getenv(k), k) for k in keys}
    missing = [k for k in keys if not os.getenv(k)]

    return {"env": vals, "missing": missing}


# ============================================================
#  CLI
# ============================================================

@app.cli.command("init-db")
def init_db_command():
    db.create_all()
    print("Initialized the database.")


@app.cli.command("migrate-allow-negative")
def migrate_allow_negative():
    """
    Migração manual para remover CHECK e incluir allow_negative.
    """
    create_sql = """
    CREATE TABLE product_new (
        id INTEGER PRIMARY KEY,
        sku VARCHAR(64) NOT NULL UNIQUE,
        name VARCHAR(200) NOT NULL,
        quantity INTEGER NOT NULL,
        min_threshold INTEGER NOT NULL,
        allow_negative INTEGER NOT NULL DEFAULT 1,
        created_at DATETIME,
        updated_at DATETIME
    );
    """

    copy_cols = """
    INSERT INTO product_new (id, sku, name, quantity, min_threshold, allow_negative, created_at, updated_at)
    SELECT id, sku, name, quantity, min_threshold,
           COALESCE(allow_negative, 1),
           created_at, updated_at
    FROM product;
    """

    with db.engine.begin() as conn:
        conn.exec_driver_sql("PRAGMA foreign_keys=OFF;")
        conn.exec_driver_sql(create_sql)
        conn.exec_driver_sql(copy_cols)
        conn.exec_driver_sql("DROP TABLE product;")
        conn.exec_driver_sql("ALTER TABLE product_new RENAME TO product;")
        conn.exec_driver_sql("PRAGMA foreign_keys=ON;")

    print("OK: migrated.")


# ============================================================
#  MAIN
# ============================================================

if __name__ == "__main__":
    with app.app_context():
        db.create_all()

    app.run(host="0.0.0.0", port=5000, debug=True)
